from django.db import models
from django.core import validators
import jwt
from datetime import datetime, timedelta
from django.conf import settings
from django.contrib.auth.models import (
	AbstractBaseUser, BaseUserManager, PermissionsMixin
)
from openpyxl import load_workbook

class UserManager(BaseUserManager):

    def create_user(self, username, email, password=None):

        if username is None:
            raise TypeError('Users must have a username.')

        if email is None:
            raise TypeError('Users must have an email address.')

        user = self.model(username=username, email=self.normalize_email(email))
        user.set_password(password)
        user.save()
        return user

    def create_superuser(self, username, email, password):

        if password is None:
            raise TypeError('Superusers must have a password.')

        user = self.create_user(username, email, password)
        user.is_superuser = True
        user.is_staff = True
        user.save()

        return user

class User(AbstractBaseUser, PermissionsMixin):

    username = models.CharField(db_index=True, max_length=255, unique=True)
    email = models.EmailField(db_index=True, unique=True)
    is_active = models.BooleanField(default=True)
    is_staff = models.BooleanField(default=False)
    created_at = models.DateTimeField(auto_now_add=True)
    updated_at = models.DateTimeField(auto_now=True)

    USERNAME_FIELD = 'email'
    REQUIRED_FIELDS = ['username']

    objects = UserManager()

    def __str__(self):
        return self.email

    @property
    def token(self):
        return self._generate_jwt_token()

    def get_full_name(self):
        return self.username

    def get_short_name(self):
        return self.username

    def _generate_jwt_token(self):
        dt = datetime.now() + timedelta(days=1)

        token = jwt.encode({
            'id': self.pk,
            'exp': int(dt.strftime('%S'))
        },settings.SECRET_KEY, algorithm='HS256')

        return token.encode('utf-8')


class NotAllData(Exception):
    pass


class ExcelFile(models.Model):
    STATUS_VALUES = [
        ('1', 'uploaded'),
        ('2', 'processed'),
        ('3', 'handled')
        ]
    excel_file = models.FileField(verbose_name='Сам файл',
                                  upload_to='files/',
                                  validators=[validators.FileExtensionValidator(allowed_extensions=('xls','xlsx'),
                                                                                message='Допускаются только файлы Excel')])
    uploaded_at = models.DateTimeField(verbose_name='Дата и ремя загрузки', auto_now_add=True)
    handled_at = models.DateTimeField(verbose_name='Дата и время окончания обработки', null=True, blank=True, default=None)
    handle_status = models.CharField(max_length=20, verbose_name='Статус обработки', choices=STATUS_VALUES,
                                         default='1')
    result = models.CharField(max_length=20, verbose_name='Результат обработки',null=True, blank=True, default=None)
    user = models.ForeignKey(User, on_delete=models.CASCADE)

    class Meta:
        verbose_name = 'Файл Excel'
        verbose_name_plural = 'Файлы Excel'

    def __str__(self):
            return self.excel_file.name

    @property
    def status(self):
        return self.get_handle_status_display()

    def delete(self, *args, **kwargs):
        self.excel_file.delete(save=False)
        super().delete(*args, **kwargs)

    def processing_file(self):
        wb = load_workbook(filename=self.excel_file.path, data_only=True, read_only=True)

        is_find = False
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            cells = iter(ws.iter_rows(min_row=1, min_col=1, max_row=1, max_col=ws.max_column))
            try:
                header_row = [(cell.value, cell.column)
                              for cell in next(cells) if cell.value == 'before' or cell.value == 'after']
            except StopIteration:
                pass

            if header_row and len(header_row) == 2:
                header_row.append(sheet)
                is_find = True
                break

        if not is_find:
            raise NotAllData('Столбцов before и after  данной таблице не существует')

        ws = wb[header_row[-1]]
        wb = None
        if header_row[0][0] == 'before':
            before_col = header_row[0][1]
            after_col = header_row[1][1]
        else:
            before_col = header_row[1][1]
            after_col = header_row[0][1]

        before_data = list()
        after_data = list()
        for i in range(2, ws.max_row + 1):
            before_item = ws.cell(row=i, column=before_col).value
            after_item = ws.cell(row=i, column=after_col).value
            if before_item:
                before_data.append(before_item)
            if after_item:
                after_data.append(after_item)

        ws = None
        if abs(len(before_data) - len(after_data)) == 1:
            if len(before_data) < len(after_data):
                res_element = f'added: {sum(after_data) - sum(before_data)}'
            else:
                res_element = f'removed: {sum(before_data) - sum(after_data)}'

            self.result = res_element
            self.handle_status = '3'
            self.handled_at = datetime.now()
            self.save()
        else:
            raise NotAllData('Разность в столбцах более чем в 1 элемент')
